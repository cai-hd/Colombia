#!/usr/bin/env python3 
# -*- coding:utf-8 _*-
""" 
@author:fjg
@license: Apache Licence 
@file: export_excel.py 
@time: 2021/04/19
@contact: fujiangong.fujg@bytedance.com
@site:  
@software: PyCharm 
"""
import pickle
import re
import time

from openpyxl import Workbook, worksheet
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter, column_index_from_string

title_font = Font(u'宋体', bold=True, size=16)
title_fill = PatternFill(patternType="solid", start_color="87CEEB")
subtitle_font = Font(u'宋体', bold=True, size=12)
subtitle_fill = PatternFill(patternType="solid", start_color="E0FFFF")
alignment = Alignment(horizontal="center", vertical="center")
error_font = Font(color="FF0000")
info_font = Font()
border = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))
with open(f'./tmp/dump-{time.strftime("%Y%m%d", time.localtime())}', 'rb') as dump:
    check_data = pickle.load(dump)


def set_data(ws: worksheet, start_column: int, start_row: int, title: list, values: list):
    max_column = max([len(x) for x in title])
    title_column = start_column
    for tit in title:
        start_column = title_column
        if len(tit) < max_column:
            ranger_cell = f"{get_column_letter(start_column)}{start_row}:{get_column_letter(max_column + start_column - 1)}{start_row}"
            for i in ws[ranger_cell]:
                for a in i:
                    a.border = border
            ws.merge_cells(ranger_cell)
        for sub_title in tit:
            cell = ws.cell(row=start_row, column=start_column)
            cell.value = sub_title
            cell.alignment = alignment
            cell.font = subtitle_font
            cell.fill = subtitle_fill
            cell.border = border
            start_column += 1
        start_row += 1
    for a in values:
        start_column = title_column
        for b in a:
            cell = ws.cell(row=start_row, column=start_column)
            if isinstance(b, tuple):
                cell.value = b[0]
                cell.font = info_font if b[1] else error_font
            else:
                cell.value = b
            cell.alignment = alignment
            cell.border = border
            start_column += 1
        start_row += 1


def set_headline(ws: worksheet, start_column: int, start_row: int, end_column: int, end_row: int, name: str):
    range_cells = f'{get_column_letter(start_column)}{start_row}:{get_column_letter(end_column)}{end_row}'
    ws.merge_cells(range_cells)
    title = ws.cell(start_row, start_column, name)
    title.fill = title_fill
    title.font = title_font
    title.alignment = alignment


def get_dimension(ws: worksheet):
    m = re.match(r'(?P<start_column>\D*)(?P<start_row>\d*):(?P<end_column>\D*)(?P<end_row>\d*)', ws.dimensions)
    start_column = column_index_from_string(m.group('start_column'))
    end_column = column_index_from_string(m.group('end_column'))
    start_row = int(m.group('start_row'))
    end_row = int(m.group('end_row'))
    return start_column, start_row, end_column, end_row


def set_dimension(ws: worksheet):
    m = re.match(r'(?P<start_column>\D*)(?P<start_row>\d*):(?P<end_column>\D*)(?P<end_row>\d*)', ws.dimensions)
    for i in range(ord(m.group('start_column')), ord(m.group('end_column')) + 1):
        ws.column_dimensions[chr(i)].width = 20
    for i in range(int(m.group('start_row')), int(m.group('end_row')) + 1):
        ws.row_dimensions[i].height = 20


def get_node_dict(key):
    node_metric = {dict(node._asdict())['node']: dict(node._asdict()) for node in
                   check_data[key]['context']['metric']['nodes']}
    node_context = {node['Hostname']: node for node in check_data[key]['context']['node']['result']}
    for node in node_metric.keys():
        node_metric[node]['memory_used'] = node_metric[node]['memory']
        del node_metric[node]['memory']
        node_metric[node]['cpu_used'] = node_metric[node]['cpu']
        del node_metric[node]['cpu']
        node_context[node]['memory_total'] = node_context[node]['memory']
        del node_context[node]['memory']
        node_context[node]['cpu_total'] = node_context[node]['cpu']
        del node_context[node]['cpu']
        node_context[node].update(node_metric[node])
        node_context[node_context[node]['InternalIP']] = node_context[node]
        del node_context[node]
    node_info = check_data[key]['node_info']
    for ip in node_info.keys():
        node_info[ip].update(node_context[ip])
    return node_info


def get_pod_dict(key):
    pod_metric = {dict(pod._asdict())['pod']: dict(pod._asdict()) for pod in
                  check_data[key]['context']['metric']['pods']}
    pod_context = {pod['name']: pod for pod in check_data[key]['context']['pod']['result']}
    for pod in pod_metric.keys():
        pod_context[pod].update(pod_metric[pod])
    return pod_context


def format_data_for_k8s(ws: worksheet, cluster: str):
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # cluster status
    head_row = 1
    head_column = 1
    # etcd status
    start_column, start_row, end_column, end_row = get_dimension(ws)
    column_column = end_column
    etcd_status_title = [['etcd status'], ['ip address', 'status', 'time']]
    etcd_status_dict = check_data[ws.title]['etcd_status']
    etcd_status_data = [[(x, True if etcd_status_dict[x]['status'] else False), etcd_status_dict[x]['status'],
                         etcd_status_dict[x]['data']] for x in etcd_status_dict.keys()]
    set_data(ws, column_column, head_row + 2, etcd_status_title, etcd_status_data)
    # node status
    start_column, start_row, end_column, end_row = get_dimension(ws)
    node_status_title = [['node status'], ['status', 'number', 'name']]
    node_status_dict = check_data[ws.title]['node_status']
    node_status_data = [[(x, node_status_dict['status']), node_status_dict['data'][x]['data'],
                         '\n'.join(node_status_dict['data'][x]['name'])] for x in node_status_dict['data'].keys()]
    set_data(ws, column_column, end_row + 2, node_status_title, node_status_data)
    # pod status
    start_column, start_row, end_column, end_row = get_dimension(ws)
    pod_status_title = [["POD"], ['status', 'number', 'name']]
    pods_status_dict = check_data[ws.title]['pods_status']
    pods_status_data = [
        [(x, pods_status_dict[x]['status']), pods_status_dict[x]['data'], '\n'.join(pods_status_dict[x]['name'])] for x
        in pods_status_dict.keys()]
    set_data(ws, column_column, end_row + 2, pod_status_title, pods_status_data)
    # 'apiserver_status', 'controller_status', 'scheduler_status'
    start_column, start_row, end_column, end_row = get_dimension(ws)
    start_column = end_column + 2
    start_row = head_row + 2
    for check_point in ['apiserver_status', 'controller_status', 'scheduler_status']:
        check_point_title = [[' '.join(check_point.split("_"))], ['ip address', 'status']]
        check_point_obj = check_data[ws.title][check_point]
        check_point_dict = dict()
        for i in check_point_obj:
            check_point_dict.update(i)
        check_point_data = [[(x, True if check_point_dict[x]['status'] else False), check_point_dict[x]['data']] for x
                            in check_point_dict.keys()]
        set_data(ws, start_column, start_row, check_point_title, check_point_data)
        start_row = start_row + len(check_point_title) + len(check_point_data) + 1
    head_end_column = head_column + max([len(x) for x in etcd_status_title]) + max([len(x) for x in check_point_title])
    set_headline(ws, head_column, head_row, head_end_column, head_row, 'cluster status')

    # network status
    start_column, start_row, end_column, end_row = get_dimension(ws)
    head_column = end_column + 2
    # coredns status
    start_column = head_column
    start_row = head_row + 2
    coredns_status_title = [['coredns status'], ['status', 'number']]
    coredns_status_dict = check_data[ws.title]['coredns_status']
    coredns_status_data = [[(x, coredns_status_dict['status']), coredns_status_dict['data'][x]] for x in
                           coredns_status_dict['data'].keys()]
    set_data(ws, start_column, start_row, coredns_status_title, coredns_status_data)
    # dns nslookup
    start_column = head_column
    start_row = start_row + len(coredns_status_title) + len(coredns_status_data) + 1
    dns_nslookup_title = [['DNS NSLOOKUP'], ['domain', 'status']]
    dns_nslookup_dict = check_data[ws.title]['dns_nslookup']
    dns_nslookup_data = [[(x, dns_nslookup_dict[x]), dns_nslookup_dict[x]] for x in dns_nslookup_dict.keys()]
    set_data(ws, start_column, start_row, dns_nslookup_title, dns_nslookup_data)
    # network
    start_column = head_column
    start_row = start_row + len(dns_nslookup_title) + len(dns_nslookup_data) + 1
    network_title = [['network route'], ['route', 'ip address']]
    network_dict = check_data[ws.title]['network']
    network_data = [[(x, network_dict[x]['status']), '\n'.join(network_dict[x]['data'])] for x in network_dict.keys()]
    set_data(ws, start_column, start_row, network_title, network_data)
    # svc cidr
    start_column, start_row, end_column, end_row = get_dimension(ws)
    start_column = end_column + 2
    start_row = head_row + 2
    svc_cidr_title = [['svc cidr'], ['cidr', 'total', 'used', 'unused']]
    svc_cidr_dict = check_data[ws.title]['svc_cidr']
    svc_cidr_data = [[(svc_cidr_dict['data']['cidr'], svc_cidr_dict['status']), svc_cidr_dict['data']['total'],
                      svc_cidr_dict['data']['used'], svc_cidr_dict['data']['unused']]]
    set_data(ws, start_column, start_row, svc_cidr_title, svc_cidr_data)
    # pod cidr
    start_row = start_row + len(svc_cidr_title) + len(svc_cidr_data) + 1
    pod_cidr_title = [['pod cidr'], ['name', 'cidr', 'total', 'used', 'unused']]
    pod_cidr_dict = check_data[ws.title]['pod_cidr']
    pod_cidr_data = [
        [(x, pod_cidr_dict[x]['status']), pod_cidr_dict[x]['data']['cidr'], pod_cidr_dict[x]['data']['total'],
         pod_cidr_dict[x]['data']['used'], pod_cidr_dict[x]['data']['unused']] for x in pod_cidr_dict.keys()]
    set_data(ws, start_column, start_row, pod_cidr_title, pod_cidr_data)
    head_end_column = head_column + max([len(x) for x in coredns_status_title]) + max([len(x) for x in pod_cidr_title])
    set_headline(ws, head_column, head_row, head_end_column, head_row, 'network status')

    # resource status
    start_column, start_row, end_column, end_row = get_dimension(ws)
    head_column = start_column
    head_row = end_row + 2
    start_row = head_row + 2
    for resource in ['cluster_quota', 'tenants_quota', 'partitions_quota']:
        resource_title = [[' '.join(resource.split("_"))], ['name', 'resource name', 'total', 'used', 'unused']]
        resource_dict = check_data[ws.title][resource]
        resource_data = [[x, (y, resource_dict[x][y]['status']), resource_dict[x][y]['data']['total'],
                          resource_dict[x][y]['data']['used'], resource_dict[x][y]['data']['unused']] for x in
                         resource_dict.keys() for y in resource_dict[x].keys()]
        if resource == 'partitions_quota':
            start_column = head_column + max([len(x) for x in resource_title]) + 1
            set_data(ws, start_column, head_row + 2, resource_title, resource_data)
        else:
            set_data(ws, head_column, start_row, resource_title, resource_data)
        start_row = start_row + len(resource_title) + len(resource_data) + 1
    head_end_column = head_column + max([len(x) for x in resource_title]) * 2
    set_headline(ws, head_column, head_row, head_end_column, head_row, 'resource status')

    # node
    start_column, start_row, end_column, end_row = get_dimension(ws)
    node_dict = get_node_dict(cluster)
    head_column = start_column
    head_row = end_row + 2
    start_row = head_row + 2
    # node info
    node_info_title = [
        ['ip', 'hostname', 'kernel', 'status', 'node load', 'container runtime', 'container process', 'docker process',
         'kubelet', 'kubeproxy', 'dns nslookup', 'disk I/O', 'network I/O', 'Z process', 'time difference']]
    node_info_data = list()
    node_resource_data = list()
    for ip in node_dict.keys():
        hostname = node_dict[ip]['Hostname']
        kernel = node_dict[ip]['kernel']
        status = (node_dict[ip]['status'], True if node_dict[ip]['status'] == "Ready" else False)
        node_load = (node_dict[ip]['nodeload']['loadaverage'], node_dict[ip]['nodeload']['check_result'])
        container_runtime = node_dict[ip]['container_runtime']
        container_process = (node_dict[ip]['containerd']['checkpass'], node_dict[ip]['containerd']['checkpass'])
        docker_process = (node_dict[ip]['docker']['dockerProcess'],
                          True if node_dict[ip]['docker']['dockerProcess'] == "active" else False)
        kubelet = (node_dict[ip]['kubelet']['porthealth'],
                   True if node_dict[ip]['kubelet']['porthealth'] == 'ok' and node_dict[ip]['kubelet'][
                       'process'] == 'active' else False)
        kubeproxy = (node_dict[ip]['kubeproxy']['porthealth'], node_dict[ip]['kubeproxy']['porthealth'])
        node_dns_data = []
        node_dns_status = True
        for domain in node_dict[ip]['dns']:
            if not domain['checkpass']:
                node_dns_status = False
                node_dns_data.append(domain['dnsname'])
        dns_nslookup = (True if node_dns_status else '\n'.join(node_dns_data), node_dns_status)
        disk_io_data = []
        disk_io_status = True
        for disk in node_dict[ip]['diskio']:
            if not disk['check_result']['isNormal']:
                disk_io_status = False
                disk_io_data.append(f"{disk['device']}:{' '.join(disk['check_result']['data'])}")
        disk_io = (True if disk_io_status else '\n'.join(disk_io_data), disk_io_status)
        network_io_data = []
        network_io_status = True
        for network in node_dict[ip]['nicio']:
            if not network['check_result']['isNormal']:
                network_io_status = False
                network_io_data.append(f"{network['device']}:{network['check_result']['data']}")
        network_io = (True if network_io_status else '\n'.join(network_io_data), network_io_status)
        z_process = (True if node_dict[ip]['zprocess']['checkpass'] else ' '.join(node_dict[ip]['zprocess']['result']),
                     node_dict[ip]['zprocess']['checkpass'])
        time_difference = (True if node_dict[ip]['ntp']['checkpass'] else node_dict[ip]['ntp']['result'],
                           node_dict[ip]['ntp']['checkpass'])
        row_data = [ip, hostname, kernel, status, node_load, container_runtime, container_process, docker_process,
                    kubelet, kubeproxy, dns_nslookup, disk_io, network_io, z_process, time_difference]
        node_info_data.append(row_data)

        # node resource data
        # cpu
        cpu_total = int(node_dict[ip]['cpu_total'])
        cpu_used = node_dict[ip]['cpu_used']
        cpu_unused = cpu_total - cpu_used
        cpu_percent = round(cpu_used / cpu_total * 100, 2)
        cpu_status = True if cpu_percent < 80 else False
        cpu_row = [ip, ('cpu', cpu_status), '', cpu_total, cpu_used, cpu_unused, f'{cpu_percent}%']
        node_resource_data.append(cpu_row)
        # memory
        memory_total = node_dict[ip]['memory_total']
        memory_used = node_dict[ip]['memory_used']
        memory_unused = f'{round(float(memory_total.rstrip("Gi")) - float(memory_used.rstrip("Gi")), 3)}Gi'
        memory_percent = round(float(memory_used.rstrip("Gi")) / float(memory_total.rstrip("Gi")) * 100, 2)
        memory_status = True if memory_percent < 80 else False
        memory_row = [ip, ('memory', memory_status), '', memory_total, memory_used, memory_unused, f'{memory_percent}%']
        node_resource_data.append(memory_row)
        # filesystem
        for fs in node_dict[ip]['diskusage']:
            fs_name = fs['Filesystem']
            fs_total = fs['Size']
            fs_used = fs['Used ']
            fs_unused = fs['Avail']
            fs_percent = fs['Use%']
            fs_status = True if int(fs_percent.rstrip('%')) < 80 else False
            fs_row = [ip, 'filesystem', (fs_name, fs_status), fs_total, fs_used, fs_unused, fs_percent]
            node_resource_data.append(fs_row)
        # docker FD
        docker_fd_total = int(node_dict[ip]['docker']['maxDockerFD'])
        docker_fd_used = int(node_dict[ip]['docker']['usedDockerFD'])
        docker_fd_unused = docker_fd_total - docker_fd_used
        docker_fs_percent = round(docker_fd_used / docker_fd_total * 100, 2)
        docker_fs_status = True if docker_fs_percent < 80 else False
        docker_fs_row = [ip, ('docker fd', docker_fs_status), '', docker_fd_total, docker_fd_used, docker_fd_unused,
                         f'{docker_fs_percent}%']
        node_resource_data.append(docker_fs_row)
        # conntrack
        conntrack_total = int(node_dict[ip]['contrack']['contrack_max'])
        conntrack_used = int(node_dict[ip]['contrack']['contrack_used'])
        conntrack_unused = conntrack_total - conntrack_used
        conntrack_percent = round(conntrack_used / conntrack_total * 100, 2)
        conntrack_status = True if conntrack_percent < 80 else False
        conntrack_row = [ip, ('conntrack', conntrack_status), '', conntrack_total, conntrack_used, conntrack_unused,
                         f'{conntrack_percent}%']
        node_resource_data.append(conntrack_row)
        # openfile 
        openfile_total = int(node_dict[ip]['openfile']['openfile_max'])
        openfile_used = int(node_dict[ip]['openfile']['openfile_used'])
        openfile_unused = openfile_total - openfile_used
        openfile_percent = round(openfile_used / openfile_total * 100, 2)
        openfile_status = True if openfile_percent < 80 else False
        openfile_row = [ip, ('openfile', openfile_status), '', openfile_total, openfile_used, openfile_unused,
                        f'{openfile_percent}%']
        node_resource_data.append(openfile_row)
        # pid
        pid_total = int(node_dict[ip]['pid']['pid_max'])
        pid_used = int(node_dict[ip]['pid']['pid_used'])
        pid_unused = pid_total - pid_used
        pid_percent = round(pid_used / pid_total * 100, 2)
        pid_status = True if pid_percent < 80 else False
        pid_row = [ip, ('pid', pid_status), '', pid_total, pid_used, pid_unused, f'{pid_percent}%']
        node_resource_data.append(pid_row)

    set_data(ws, start_column, start_row, node_info_title, node_info_data)
    # node resource quota
    start_column, start_row, end_column, end_row = get_dimension(ws)
    start_row = end_row + 2
    node_resource_title = [['node resource quota'], ['ip', 'resource', 'name', 'total', 'used', 'unused', 'percent']]
    set_data(ws, start_column, start_row, node_resource_title, node_resource_data)
    head_end_column = head_column + max([len(x) for x in node_info_title]) - 1
    set_headline(ws, head_column, head_row, head_end_column, head_row, 'node info')

    # pod info
    start_column, start_row, end_column, end_row = get_dimension(ws)
    pod_dict = get_pod_dict(cluster)
    head_column = start_column
    head_row = end_row + 2
    start_row = head_row + 2
    pod_info_title = [
        ['name', 'ns', 'status', 'ip', 'node', 'restart', 'cpu used', 'cpu requests', 'cpu limits', 'memory used',
         'memory requests', 'memory limits']]
    pod_info_data = []
    for pod in pod_dict.keys():
        cpu_status = True if pod_dict[pod]['cpu'] < pod_dict[pod]['cpu_limits'] * 0.8 else False
        mem_used = float(pod_dict[pod]['memory'].rstrip("Mi")) / 1024
        mem_limits = float(pod_dict[pod]['memory_limits'].rstrip("Gi"))
        memory_status = True if mem_used < mem_limits * 0.8 else False
        restart_status = True if pod_dict[pod]['restart'] < 20 else False
        status_status = True if pod_dict[pod]['status'] in ['Running', 'Succeeded'] else False
        pod_info_row = [pod, pod_dict[pod]['ns'], (pod_dict[pod]['status'], status_status), pod_dict[pod]['ip'],
                        pod_dict[pod]['host'], (pod_dict[pod]['restart'], restart_status),
                        (pod_dict[pod]['cpu'], cpu_status), (pod_dict[pod]['cpu_requests'], cpu_status),
                        (pod_dict[pod]['cpu_limits'], cpu_status), (pod_dict[pod]['memory'], memory_status),
                        (pod_dict[pod]['memory_requests'], memory_status),
                        (pod_dict[pod]['memory_limits'], memory_status)]

        pod_info_data.append(pod_info_row)
    set_data(ws, start_column, start_row, pod_info_title, pod_info_data)
    head_end_column = max([len(x) for x in pod_info_title])
    set_headline(ws, head_column, head_row, head_end_column, head_row, 'pod info')


def format_other_data(ws: worksheet):
    # license
    license_title = [['license'], ['resource', 'remain', 'total/start', 'used/end']]
    license_dict = check_data['license']
    license_data = list()
    for key in license_dict.keys():
        if key == 'day':
            row = [(key, license_dict[key]['status']), license_dict[key]['unused'], license_dict[key]['start'],
                   license_dict[key]['end']]
        else:
            row = [(key, license_dict[key]['status']), license_dict[key]['unused'], license_dict[key]['total'],
                   license_dict[key]['used']]
        license_data.append(row)
    set_data(ws, 1, 1, license_title, license_data)

    # volumes
    volumes_title = [['volumes'], ['cluster', 'name', 'brick']]
    volumes_dict = check_data['volumes_status']
    volumes_data = [[x, (y, volumes_dict[x][y]['status']), '\n'.join(volumes_dict[x][y]['data'])] for x in
                    volumes_dict.keys() for y in volumes_dict[x].keys()]
    start_column, start_row, end_column, end_row = get_dimension(ws)
    start_row = end_row + 2
    set_data(ws, start_column, start_row, volumes_title, volumes_data)


def main():
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    for cluster in check_data.keys():
        if cluster not in ['license', 'volumes_status']:
            ws = wb.create_sheet(cluster)
            format_data_for_k8s(ws, cluster)
            set_dimension(ws)
    ws = wb.create_sheet("others")
    format_other_data(ws)
    set_dimension(ws)
    wb.save(f'./tmp/{time.strftime("%Y%m%d", time.localtime())}.xlsx')


if __name__ == '__main__':
    main()
